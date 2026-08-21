#!/usr/bin/env python3
"""Genera deploy/comandos-operador.xlsx para copiar y pegar en cada máquina."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = "deploy/comandos-operador.xlsx"

NAVY = "0D1B2A"
BLUE = "1263F5"
BLUE_SOFT = "EAF1FF"
INK = "0D1B2A"
MUTED = "5B6B80"
LINE = "C8D4E6"
OK = "E5F6EC"
WARN = "FDF3E2"
HEAD_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color=NAVY, size=16)
SUB_FONT = Font(name="Calibri", italic=True, color=MUTED, size=11)
CELL_FONT = Font(name="Calibri", size=11)
MONO = Font(name="Consolas", size=10)
HEAD_FILL = PatternFill("solid", fgColor=BLUE)
TITLE_FILL = PatternFill("solid", fgColor=BLUE_SOFT)
WARN_FILL = PatternFill("solid", fgColor=WARN)
OK_FILL = PatternFill("solid", fgColor=OK)
THIN = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)
WRAP = Alignment(wrap_text=True, vertical="center")
TOP = Alignment(wrap_text=True, vertical="top")


def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN


def paint(ws, r, c, value, font=CELL_FONT, fill=None, align=WRAP):
    cell = ws.cell(r, c, value)
    cell.font = font
    cell.alignment = align
    cell.border = THIN
    if fill:
        cell.fill = fill
    return cell


def widths(ws, sizes):
    for i, size in enumerate(sizes, 1):
        ws.column_dimensions[get_column_letter(i)].width = size


def build():
    wb = Workbook()

    # ----- Config -----
    cfg = wb.active
    cfg.title = "1. Config"
    cfg.sheet_properties.tabColor = BLUE
    cfg.merge_cells("A1:C1")
    paint(cfg, 1, 1, "ARGOS · cambiá estos valores y el resto de hojas arma los comandos", TITLE_FONT, TITLE_FILL)
    cfg["B1"].border = THIN
    cfg["C1"].border = THIN
    cfg.row_dimensions[1].height = 28

    headers = ["Clave", "Valor (editá acá)", "Para qué"]
    for i, h in enumerate(headers, 1):
        paint(cfg, 3, i, h)
    style_header(cfg, 3, 3)

    rows = [
        ("repo", "https://github.com/kodeskia-byte/ARGOS.git", "git clone en un servidor nuevo"),
        ("dir", "ARGOS", "Carpeta del repo en el servidor (cd ARGOS)"),
        ("controller_url", "http://127.0.0.1:8080", "Con túnel inverso el runner ve el collector en localhost"),
        ("flow", "flows/example.yaml", "YAML del journey"),
        ("data", "", "CSV opcional. Ej: flows/data.example.csv. Vacío = sin --data"),
        ("ramp", "5@1m,15@3m,30@5m", "Rampa: usuarios@tiempo. Una sola corrida"),
        ("users", "20", "Solo si no usás rampa (escalón fijo)"),
        ("duration", "10m", "Solo si no usás rampa"),
        ("abort_error", "0.4", "Corta si el error llega al 40%"),
        ("abort_cpu", "90", "Corta si la CPU del GENERADOR llega a 90%"),
        ("lite", "si", "si = --lite (capacidad). no = Chromium de usuario real"),
        ("ip_pc", "127.0.0.1", "IP de la PC donde corre el collector (para el túnel)"),
        ("puerto", "8080", "Puerto del Live Room"),
        ("user_ssh", "root", "Usuario SSH de los generadores"),
    ]
    for i, (k, v, w) in enumerate(rows, 4):
        paint(cfg, i, 1, k, Font(name="Consolas", bold=True, size=11))
        paint(cfg, i, 2, v, MONO, WARN_FILL)
        paint(cfg, i, 3, w)
        cfg.row_dimensions[i].height = 22
    widths(cfg, [22, 55, 70])
    cfg.freeze_panes = "A4"
    cfg.row_dimensions[3].height = 22
    paint(
        cfg,
        19,
        1,
        "El collector corre en tu PC con ./venv/bin/python (el python del sistema es 2.7 y se cae). "
        "El túnel inverso deja el 8080 de la PC en el 8080 de cada servidor. "
        "Por eso controller_url es http://127.0.0.1:8080 cuando lanzás desde el generador.",
        SUB_FONT,
        TITLE_FILL,
        TOP,
    )
    cfg.merge_cells("A19:C19")
    cfg.row_dimensions[19].height = 48
    cfg["B19"].border = THIN
    cfg["C19"].border = THIN

    # ----- Generadores -----
    gens = wb.create_sheet("2. Generadores")
    gens.sheet_properties.tabColor = BLUE
    gens.merge_cells("A1:E1")
    paint(
        gens,
        1,
        1,
        "Una fila = un SERVIDOR (generador). Las sondas probe-01, probe-02… las crea el runner adentro. "
        "El instance_id tiene que ser distinto: si dos máquinas dicen gen-01, el informe se mezcla.",
        TITLE_FONT,
        TITLE_FILL,
    )
    for col in range(2, 6):
        gens.cell(1, col).border = THIN
        gens.cell(1, col).fill = TITLE_FILL
    gens.row_dimensions[1].height = 40
    for i, h in enumerate(["#", "instance_id", "ssh_destino (user@IP)", "IP sola", "Notas"], 1):
        paint(gens, 3, i, h)
    style_header(gens, 3, 5)

    # gen-01 with known IP from operations; rest placeholders
    known = ["216.185.51.83"] + [""] * 9
    for i in range(1, 11):
        r = 3 + i
        paint(gens, r, 1, i)
        paint(gens, r, 2, f"gen-{i:02d}", Font(name="Consolas", bold=True, size=12), OK_FILL)
        ip = known[i - 1] or f"IP_GEN_{i:02d}"
        # ssh_destino formula from Config user_ssh
        paint(gens, r, 3, None, MONO, WARN_FILL)
        gens.cell(r, 3).value = f'=CONCATENATE(\'1. Config\'!$B$17,"@",D{r})'
        paint(gens, r, 4, ip, MONO, WARN_FILL)
        paint(gens, r, 5, "Completá la IP. Dejá la fila si no usás este servidor." if i > 1 else "Servidor que ya usamos")
        gens.row_dimensions[r].height = 22
    # Fix user_ssh cell: Config rows start at 4, user_ssh is the 14th data row = row 17. Yes B17.

    widths(gens, [6, 16, 32, 22, 62])
    gens.freeze_panes = "A4"
    paint(
        gens,
        15,
        1,
        "Con 1 servidor solo usás la fila gen-01. Con 10, las 10. fleet.sh lee deploy/hosts.txt "
        "(una línea user@IP por máquina) y pone gen-01… gen-10 solo.",
        SUB_FONT,
        TITLE_FILL,
        TOP,
    )
    gens.merge_cells("A15:E15")
    for col in range(2, 6):
        gens.cell(15, col).border = THIN
        gens.cell(15, col).fill = PatternFill("solid", fgColor=BLUE_SOFT)
    gens.row_dimensions[15].height = 36

    # ----- PC -----
    pc = wb.create_sheet("3. En la PC")
    pc.sheet_properties.tabColor = "0F9D58"
    pc.merge_cells("A1:C1")
    paint(pc, 1, 1, "Esto se corre en TU notebook, no en el servidor de carga", TITLE_FONT, TITLE_FILL)
    pc["B1"].border = THIN
    pc["C1"].border = THIN
    pc.row_dimensions[1].height = 28
    for i, h in enumerate(["Paso", "Dónde", "Comando (copiá la celda)"], 1):
        paint(pc, 3, i, h)
    style_header(pc, 3, 3)

    pc_rows = [
        (
            "1. Collector (dejalo abierto)",
            "PC · terminal 1",
            '="cd ~/Música/ARGOS && ./venv/bin/python -m argos.controller.server --host 127.0.0.1 --port "&\'1. Config\'!B16',
        ),
        (
            "2. Túnel inverso (dejalo abierto)",
            "PC · terminal 2",
            '="ssh -N -o ServerAliveInterval=30 -o ExitOnForwardFailure=yes -R "&\'1. Config\'!B16&":127.0.0.1:"&\'1. Config\'!B16&" "&\'2. Generadores\'!C4',
        ),
        (
            "3. Dashboard",
            "Navegador en la PC",
            '="http://127.0.0.1:"&\'1. Config\'!B16',
        ),
        (
            "4. Si el puerto 8080 está ocupado",
            "PC",
            "fuser -k 8080/tcp",
        ),
        (
            "5. NO uses python a secas",
            "PC",
            "# el python del sistema es 2.7. Siempre ./venv/bin/python",
        ),
        (
            "6. git pull en la PC (código nuevo)",
            "PC · carpeta ARGOS",
            "cd ~/Música/ARGOS && git pull",
        ),
        (
            "7. Reiniciar collector después de git pull",
            "PC · terminal 1",
            "Ctrl+C y volvé a lanzar el paso 1. Python no recarga el HTML solo.",
        ),
    ]
    for i, (paso, donde, cmd) in enumerate(pc_rows, 4):
        paint(pc, i, 1, paso)
        paint(pc, i, 2, donde)
        cell = paint(pc, i, 3, cmd if not str(cmd).startswith("=") else None, MONO, OK_FILL, TOP)
        if str(cmd).startswith("="):
            cell.value = cmd
            cell.font = MONO
            cell.fill = OK_FILL
        pc.row_dimensions[i].height = 36
    widths(pc, [42, 22, 100])
    pc.freeze_panes = "A4"

    # ----- SSH -----
    ssh = wb.create_sheet("4. SSH por generador")
    ssh.sheet_properties.tabColor = BLUE
    ssh.merge_cells("A1:D1")
    paint(ssh, 1, 1, "Conectarte a cada servidor. Una fila = un generador. El instance_id ya va distinto.", TITLE_FONT, TITLE_FILL)
    for col in range(2, 5):
        ssh.cell(1, col).border = THIN
        ssh.cell(1, col).fill = TITLE_FILL
    ssh.row_dimensions[1].height = 28
    for i, h in enumerate(["instance_id", "Conectar", "Una vez adentro", "Salir"], 1):
        paint(ssh, 3, i, h)
    style_header(ssh, 3, 4)
    for i in range(10):
        r = 4 + i
        src = 4 + i  # Generadores data starts row 4
        paint(ssh, r, 1, f"='2. Generadores'!B{src}", Font(name="Consolas", bold=True, size=12), OK_FILL)
        paint(ssh, r, 2, f'="ssh "&\'2. Generadores\'!C{src}', MONO, WARN_FILL, TOP)
        paint(ssh, r, 3, f'="cd "&\'1. Config\'!B5', MONO)
        paint(ssh, r, 4, "exit")
        ssh.row_dimensions[r].height = 22
    widths(ssh, [16, 42, 18, 12])
    ssh.freeze_panes = "A4"

    # ----- Instalar -----
    inst = wb.create_sheet("5. Instalar y actualizar")
    inst.sheet_properties.tabColor = "E8A33D"
    inst.merge_cells("A1:C1")
    paint(inst, 1, 1, "Correr en CADA generador (después de ssh). Primera vez = instalar. Después = git pull.", TITLE_FONT, TITLE_FILL)
    inst["B1"].border = THIN
    inst["C1"].border = THIN
    inst.row_dimensions[1].height = 28
    for i, h in enumerate(["Cuándo", "Qué", "Comando"], 1):
        paint(inst, 3, i, h)
    style_header(inst, 3, 3)
    inst_rows = [
        (
            "Servidor nuevo",
            "Clonar el repo",
            '="git clone "&\'1. Config\'!B4&" && cd "&\'1. Config\'!B5',
        ),
        (
            "Servidor nuevo",
            "Instalar generador (venv, Chromium, fuentes PDF)",
            '="cd "&\'1. Config\'!B5&" && ./deploy/setup.sh"',
        ),
        (
            "Solo la PC del collector, si el collector va en un VPS",
            "Instalar collector como servicio",
            '="cd "&\'1. Config\'!B5&" && ./deploy/setup.sh --with-collector"',
        ),
        (
            "Antes de cada prueba",
            "Bajar el código nuevo",
            '="cd "&\'1. Config\'!B5&" && git pull"',
        ),
        (
            "Verificar que ARGOS responde",
            "Tiene que imprimir ok",
            '="cd "&\'1. Config\'!B5&" && ./venv/bin/python -c ""import argos; print(\'ok\')"""',
        ),
        (
            "Ver la hora (NTP)",
            "Las 10 máquinas tienen que coincidir",
            "date",
        ),
        (
            "Humo: 1 sonda 30 s (este generador)",
            "Reemplazá gen-01 por el instance_id de ESTA máquina",
            '="cd "&\'1. Config\'!B5&" && ./venv/bin/python runner.py --users 1 --duration 30s --flow "&\'1. Config\'!B7&" --controller-url "&\'1. Config\'!B6&" --instance-id gen-01"',
        ),
    ]
    for i, (a, b, c) in enumerate(inst_rows, 4):
        paint(inst, i, 1, a)
        paint(inst, i, 2, b)
        cell = paint(inst, i, 3, None, MONO, OK_FILL, TOP)
        cell.value = c
        inst.row_dimensions[i].height = 40
    widths(inst, [38, 48, 110])
    inst.freeze_panes = "A4"

    def lite_bit():
        return 'IF(\'1. Config\'!B14="si"," --lite","")'

    def data_bit():
        return 'IF(\'1. Config\'!B8="",""," --data "&\'1. Config\'!B8)'

    def ramp_cmd(instance_cell):
        c = "'1. Config'"
        return (
            f'="cd "&{c}!B5&" && nohup ./venv/bin/python runner.py'
            f' --flow "&{c}!B7&"'
            f' --ramp "&{c}!B9&"'
            f' --abort-error "&{c}!B12&"'
            f' --abort-cpu "&{c}!B13&"'
            f' --controller-url "&{c}!B6&"'
            f' --instance-id "&{instance_cell}'
            f'&{data_bit()}&{lite_bit()}&" > carga.log 2>&1 &"'
        )

    def fixed_cmd(instance_cell):
        c = "'1. Config'"
        return (
            f'="cd "&{c}!B5&" && nohup ./venv/bin/python runner.py'
            f' --flow "&{c}!B7&"'
            f' --users "&{c}!B10&"'
            f' --duration "&{c}!B11&"'
            f' --controller-url "&{c}!B6&"'
            f' --instance-id "&{instance_cell}'
            f'&{data_bit()}&{lite_bit()}&" > carga.log 2>&1 &"'
        )

    # ----- Lanzar rampa -----
    ramp = wb.create_sheet("6. Lanzar RAMPA")
    ramp.sheet_properties.tabColor = "DC3545"
    ramp.merge_cells("A1:D1")
    paint(
        ramp,
        1,
        1,
        "PEGÁ UN COMANDO POR SERVIDOR. Cada uno tiene --instance-id distinto. nohup sobrevive si se corta el SSH. "
        "La rampa se edita en 1. Config (celda ramp).",
        TITLE_FONT,
        TITLE_FILL,
    )
    for col in range(2, 5):
        ramp.cell(1, col).border = THIN
        ramp.cell(1, col).fill = TITLE_FILL
    ramp.row_dimensions[1].height = 40
    for i, h in enumerate(["instance_id", "ssh_destino", "Comando a pegar EN ESE servidor", "Ver log"], 1):
        paint(ramp, 3, i, h)
    style_header(ramp, 3, 4)
    for i in range(10):
        r = 4 + i
        src = 4 + i
        paint(ramp, r, 1, f"='2. Generadores'!B{src}", Font(name="Consolas", bold=True, size=12), OK_FILL)
        paint(ramp, r, 2, f"='2. Generadores'!C{src}", MONO)
        paint(ramp, r, 3, ramp_cmd(f"'2. Generadores'!B{src}"), MONO, WARN_FILL, TOP)
        paint(ramp, r, 4, "tail -f carga.log", MONO)
        ramp.row_dimensions[r].height = 48
    widths(ramp, [16, 32, 120, 22])
    ramp.freeze_panes = "A4"
    paint(
        ramp,
        15,
        1,
        "Orden: 1) collector en la PC  2) túnel  3) git pull en cada gen  4) pegar el comando de esta hoja. "
        "Si usás un solo servidor, solo la fila gen-01. Con 10, las 10, lo más junto que puedas "
        "(o usá la hoja 8. Flota).",
        SUB_FONT,
        TITLE_FILL,
        TOP,
    )
    ramp.merge_cells("A15:D15")
    for col in range(2, 5):
        ramp.cell(15, col).border = THIN
        ramp.cell(15, col).fill = PatternFill("solid", fgColor=BLUE_SOFT)
    ramp.row_dimensions[15].height = 40

    # ----- Lanzar fijo -----
    fijo = wb.create_sheet("7. Lanzar USERS fijo")
    fijo.sheet_properties.tabColor = MUTED
    fijo.merge_cells("A1:C1")
    paint(
        fijo,
        1,
        1,
        "Escalón clásico (--users + --duration). Sin rampa. Editá users y duration en 1. Config.",
        TITLE_FONT,
        TITLE_FILL,
    )
    fijo["B1"].border = THIN
    fijo["C1"].border = THIN
    fijo.row_dimensions[1].height = 28
    for i, h in enumerate(["instance_id", "ssh_destino", "Comando a pegar EN ESE servidor"], 1):
        paint(fijo, 3, i, h)
    style_header(fijo, 3, 3)
    for i in range(10):
        r = 4 + i
        src = 4 + i
        paint(fijo, r, 1, f"='2. Generadores'!B{src}", Font(name="Consolas", bold=True, size=12), OK_FILL)
        paint(fijo, r, 2, f"='2. Generadores'!C{src}", MONO)
        paint(fijo, r, 3, fixed_cmd(f"'2. Generadores'!B{src}"), MONO, WARN_FILL, TOP)
        fijo.row_dimensions[r].height = 44
    widths(fijo, [16, 32, 130])
    fijo.freeze_panes = "A4"

    # ----- Flota -----
    fleet = wb.create_sheet("8. Flota (10 de un tiro)")
    fleet.sheet_properties.tabColor = BLUE
    fleet.merge_cells("A1:B1")
    paint(fleet, 1, 1, "Desde la PC, si ya tenés deploy/hosts.txt. Un run_id compartido. No pegues 10 ssh a mano.", TITLE_FONT, TITLE_FILL)
    fleet["B1"].border = THIN
    fleet.row_dimensions[1].height = 28
    for i, h in enumerate(["Paso", "Comando"], 1):
        paint(fleet, 3, i, h)
    style_header(fleet, 3, 2)
    fleet_rows = [
        (
            "Armar hosts.txt (una vez)",
            "cp deploy/hosts.example.txt deploy/hosts.txt   # y poné user@IP, una por línea",
        ),
        (
            "Verificar SSH + Python + NTP",
            '="cd ~/Música/ARGOS && ./deploy/fleet.sh --hosts deploy/hosts.txt --check"',
        ),
        (
            "Rampa en todos (misma hora NTP)",
            '="cd ~/Música/ARGOS && export ARGOS_CONTROLLER_URL="&\'1. Config\'!B6&" && ./deploy/fleet.sh --hosts deploy/hosts.txt --pull --at 18:30:00 --flow "&\'1. Config\'!B7&" --ramp "&\'1. Config\'!B9&" --abort-error "&\'1. Config\'!B12&" --abort-cpu "&\'1. Config\'!B13&IF(\'1. Config\'!B14="si"," --lite","")&IF(\'1. Config\'!B8="",""," --data "&\'1. Config\'!B8)',
        ),
        (
            "Cortar toda la flota",
            "cd ~/Música/ARGOS && ./deploy/fleet.sh --hosts deploy/hosts.txt --stop",
        ),
    ]
    for i, (a, b) in enumerate(fleet_rows, 4):
        paint(fleet, i, 1, a)
        cell = paint(fleet, i, 2, b if not str(b).startswith("=") else None, MONO, OK_FILL, TOP)
        if str(b).startswith("="):
            cell.value = b
            cell.font = MONO
            cell.fill = OK_FILL
        fleet.row_dimensions[i].height = 48
    widths(fleet, [36, 130])
    fleet.freeze_panes = "A4"

    # ----- Cortar / logs -----
    stop = wb.create_sheet("9. Cortar y logs")
    stop.sheet_properties.tabColor = "DC3545"
    stop.merge_cells("A1:C1")
    paint(stop, 1, 1, "Emergencia y seguimiento. El instance_id no se usa para matar: se mata el runner de ESA máquina.", TITLE_FONT, TITLE_FILL)
    stop["B1"].border = THIN
    stop["C1"].border = THIN
    for i, h in enumerate(["Dónde", "Qué", "Comando"], 1):
        paint(stop, 3, i, h)
    style_header(stop, 3, 3)
    stop_rows = [
        ("Ese generador (ya adentro por SSH)", "Cortar la carga", "pkill -f runner.py"),
        ("Ese generador", "Ver si sigue vivo", "pgrep -af runner.py"),
        ("Ese generador", "Log en vivo", "tail -f carga.log"),
        ("Flota desde la PC", "Cortar los 10", "cd ~/Música/ARGOS && ./deploy/fleet.sh --hosts deploy/hosts.txt --stop"),
        ("PC", "Collector como servicio", "sudo systemctl restart argos-collector"),
        ("PC", "Log del collector servicio", "sudo journalctl -u argos-collector -f"),
        ("Navegador", "Live Room", '="http://127.0.0.1:"&\'1. Config\'!B16'),
        ("Navegador", "Informe de un generador", '="http://127.0.0.1:"&\'1. Config\'!B16&"/informe?instance=gen-01&run=PEGÁ_EL_RUN_ID"'),
    ]
    for i, (a, b, c) in enumerate(stop_rows, 4):
        paint(stop, i, 1, a)
        paint(stop, i, 2, b)
        cell = paint(stop, i, 3, c if not str(c).startswith("=") else None, MONO, WARN_FILL, TOP)
        if str(c).startswith("="):
            cell.value = c
            cell.font = MONO
            cell.fill = WARN_FILL
        stop.row_dimensions[i].height = 32
    widths(stop, [38, 32, 100])
    stop.freeze_panes = "A4"

    # ----- Cómo leer -----
    how = wb.create_sheet("0. Cómo usar este Excel", 0)
    how.sheet_properties.tabColor = NAVY
    how.merge_cells("A1:B1")
    paint(how, 1, 1, "ARGOS · comandos del operador", TITLE_FONT, TITLE_FILL)
    how["B1"].border = THIN
    how.row_dimensions[1].height = 28
    how.merge_cells("A2:B2")
    paint(
        how,
        2,
        1,
        "No memorices comandos. Completá 1. Config y 2. Generadores. Después copiá la celda verde o amarilla.",
        SUB_FONT,
        TITLE_FILL,
    )
    how["B2"].border = THIN
    how.row_dimensions[2].height = 28
    for i, h in enumerate(["Hoja", "Para qué"], 1):
        paint(how, 4, i, h)
    style_header(how, 4, 2)
    guide = [
        ("1. Config", "URL del collector, YAML, rampa, --lite. Se edita UNA vez por prueba."),
        ("2. Generadores", "IP de cada servidor. gen-01 ya tiene 216.185.51.83. Completá el resto."),
        ("3. En la PC", "Collector + túnel inverso. Sin esto el runner no aparece en el dashboard."),
        ("4. SSH por generador", "ssh root@IP de cada máquina. instance_id distinto por fila."),
        ("5. Instalar y actualizar", "clone, setup.sh, git pull, humo de 1 sonda."),
        ("6. Lanzar RAMPA", "EL comando de hoy: 10 filas, --instance-id gen-01 … gen-10."),
        ("7. Lanzar USERS fijo", "Si no querés rampa: --users y --duration."),
        ("8. Flota", "Los 10 de un tiro con fleet.sh y un run_id compartido."),
        ("9. Cortar y logs", "pkill, tail -f carga.log, informe."),
    ]
    for i, (a, b) in enumerate(guide, 5):
        paint(how, i, 1, a, Font(name="Calibri", bold=True, size=11), OK_FILL)
        paint(how, i, 2, b)
        how.row_dimensions[i].height = 22
    widths(how, [24, 100])
    how.merge_cells("A15:B15")
    paint(
        how,
        15,
        1,
        "Sonda ≠ generador. Generador = una máquina (gen-01). Sonda = un Chromium adentro (probe-01). "
        "Vos lanzás UN comando por generador. ARGOS abre las sondas solo. "
        "Celdas amarillas = las editás. Celdas verdes = las copiás.",
        SUB_FONT,
        WARN_FILL,
        TOP,
    )
    how["B15"].border = THIN
    how.row_dimensions[15].height = 48
    how.freeze_panes = "A5"

    # print titles
    for ws in wb.worksheets:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_view.showGridLines = False
        ws.print_title_rows = "1:3"

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    build()
